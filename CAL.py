import openpyxl
import numpy as np
from datetime import datetime
import datetime
import pandas as pd
import pymongo
#import Convert_toSQL as cts



# excel_file1 = openpyxl.load_workbook(r"C:\Users\AMIT\OneDrive\Desktop\Forecast Report Format.xlsx")
# excel_sheet1 = excel_file1["Day on Day FC"]

# excel_file2 = openpyxl.load_workbook(r"C:\Users\AMIT\OneDrive\Desktop\History and Forecast Report-20230206.xlsx")
# excel_sheet2 = excel_file2['History and Forecast Report']

# # Code for manually adding group confirm number
# group_confirm = int(input("Enter group confirm-"))
# excel_sheet2.cell(row=5, column=11).value = group_confirm
# excel_file2.save(r"C:\Users\AMIT\OneDrive\Desktop\History and Forecast Report-20230206.xlsx")

def perform(excel_file1,excel_sheet1, excel_sheet2,excel_sheet3,excel_sheet4):
    occupancy_date = []
    for row in excel_sheet1.iter_rows(min_row=1, min_col=6, max_col=6, values_only=True):
        occupancy_date.append(row[0])
    
    date_array = []
    for row in excel_sheet2.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
        date_array.append(row[0]) 

    start = 0
    end = 0

    for i in range(len(date_array)):
        if type(date_array[i]) == datetime.datetime:
            start = i
            break

    for i in range(len(date_array)-1, -1, -1):
        if type(date_array[i]) == datetime.datetime:
            end = i
            break

    start1 = start
    start2 = start + 3

    for i in range(len(occupancy_date)):
        if type(occupancy_date[i]) == datetime.datetime:
            index = start1 if occupancy_date[i] == date_array[start1] else np.searchsorted(date_array[start2:end+1], occupancy_date[i]) + start2
            if index >= start and date_array[index] == occupancy_date[i]:
                print(f"The given date array contains the date {occupancy_date[i]}")
                rs_fit = excel_sheet2.cell(row=index+1, column=9).value + excel_sheet2.cell(row=index+1, column=10).value
                rs_groups = excel_sheet2.cell(row=index+1, column=11).value + excel_sheet2.cell(row=index+1, column=12).value
                rs_CH = excel_sheet2.cell(row=index+1, column=6).value + excel_sheet2.cell(row=index+1, column=7).value
                #pk_up_fit  = excel_sheet1.cell(row=1,column=11).value -  excel_sheet1.cell(row=1,column=8).value

                excel_sheet1.cell(row=i+1, column=8).value = rs_fit
                excel_sheet1.cell(row=i+1, column=9).value = rs_groups
                excel_sheet1.cell(row=i+1, column=10).value = rs_CH
                
                excel_sheet1.cell(row=i+1, column=12).value = rs_fit
                excel_sheet1.cell(row=i+1, column=13).value = rs_groups
                excel_sheet1.cell(row=i+1, column=14).value = rs_CH
                #excel_sheet1.cell(row=i+1, column=)

    #lyrs  fit AG column 33
    connection_uri = "mongodb+srv://annu21312:6dPsrXPfhm19YxXl@hello.hes3iy5.mongodb.net/"
    client = pymongo.MongoClient(connection_uri, serverSelectionTimeoutMS=30000)
    database_name = "Revenue_Forecasting"
    db = client[database_name]
    collection2 = db["History_Fore"]
    cursor2 = collection2.find({})
    df = pd.DataFrame(list(cursor2))
    saleable = df.loc[7:len(df),'Unnamed: 6'].tolist()
    for j in range(0,len(saleable)):
        excel_sheet1.cell(row=j+5, column=33).value = saleable[j]

    collection1 = db["History"]
    cursor1 = collection1.find({})
    df2 = pd.DataFrame(list(cursor1))
    room_sold = df2.loc[3:55,'Unnamed: 2'].tolist()
    for j in range(len(room_sold)):
        excel_sheet1.cell(row=317+j, column=33).value = room_sold[j]


    


    #lyrs groups AH column 34
    for j in range(0,369):
        excel_sheet1.cell(row=j+5, column=34).value = 0

    #lyrs comp/house AI column 35
    house_use = df.loc[8:len(df),'Unnamed: 15'].tolist()
    for j in range(0,len(house_use)):
        excel_sheet1.cell(row=j+5, column=35).value = house_use[j]
    
    house_use1 = df2.loc[3:55,'Unnamed: 6'].tolist()
    for j in range(len(house_use1)):
        excel_sheet1.cell(row  = 317+j,column = 35).value = house_use1[j]

    

    #ly actual fit AK column 37
    fit3 = df.loc[7:len(df),'Unnamed: 19'].tolist()
    for j in range(0,len(fit3)):
        excel_sheet1.cell(row=j+5, column=37).value = fit3[j]
   
    fit = df2.loc[3:55,'Unnamed: 8'].tolist()
    fit2 =df2.loc[3:55,'Unnamed: 9'].tolist()
    for j in range(len(fit)):
        excel_sheet1.cell(row=317+j, column=37).value = fit[j] + fit2[j]

    
    
    #ly actual groups  AL COLUMN 38
    for j in range(0,369):
        excel_sheet1.cell(row=j+5, column=38).value = 0

    #ly actual comp/house  AM Column 39
    for j in range(0,369):
        excel_sheet1.cell(row=j+5, column=39).value = 0

    
    

    #for 2nd file 6 Feb - 9 Sep
    #converting all dates to month for getting data for specific month
    collection5 = db["Revenue"]
    cursor5 = collection5.find({})
    monthly_data1 = pd.DataFrame(list(cursor5))
    #monthly_data1 = pd.read_excel('revenue.xlsx')
    monthly_data1['Business Date'] = pd.to_datetime(monthly_data1['Business Date'])
    monthly_data1['Month'] = monthly_data1['Business Date'].dt.strftime('%Y-%m')
    
     #finding sum of revenue in all months
    monthly_total_revenue1 = monthly_data1.groupby('Month')['Room Revenue'].sum().reset_index()
    monthly_ARR =  monthly_data1.groupby('Month')['ARR'].sum().reset_index()
    Individual_ARR = monthly_data1.groupby('Month')['Individual ARR'].sum().reset_index()
    Group_Confirmed_Arr = monthly_data1.groupby('Month')['Confirmed Group ARR'].sum().reset_index()



    #Revenue Summary excel sheet 3
    collection3 = db["Covid"]
    cursor3 = collection3.find({"Date": {"$type": ["date", "null"]}})
    monthly_data = pd.DataFrame(list(cursor3))
    
   

    value = "1"
    monthly_data.fillna(value, inplace=True)
    #converting all dates to month for getting data for specific month
    #monthly_data = pd.read_excel(excel_file_path)
    monthly_data['Date'] = pd.to_datetime(monthly_data['Date'])
    monthly_data['Month'] = monthly_data['Date'].dt.strftime('%Y-%m')
    #finding sum of revenue in all months
    monthly_total_revenue = monthly_data.groupby('Month')['Revenue'].sum().reset_index()
    df23 = monthly_data.groupby('Month')['Unnamed: 27'].sum().reset_index()
    df24 = monthly_data.groupby('Month')['Unnamed: 19'].sum().reset_index()


    

    for i in  range(6,17):
        excel_sheet3.cell(row=i, column=7).value = monthly_total_revenue.loc[i,'Revenue']
        excel_sheet3.cell(row=i, column=5).value = df23.loc[i,'Unnamed: 27']
        excel_sheet3.cell(row=i, column=6).value = df24.loc[i,'Unnamed: 19']
        

    
    for i in range(6,14):
         excel_sheet3.cell(row=i, column=4).value = monthly_total_revenue1.loc[i - 6,'Room Revenue']
    j =3
    for i in range(6,14):
       
       
        excel_sheet4.cell(row=j, column=13).value = monthly_ARR.loc[i - 6,'ARR']
        excel_sheet4.cell(row=j+1, column=13).value = Individual_ARR.loc[i - 6,'Individual ARR']
        excel_sheet4.cell(row=j+2, column=13).value = Group_Confirmed_Arr.loc[i - 6,'Confirmed Group ARR']

        j+=4

    excel_file1.save(r"D:\Arjun_Sir-IP-master\Upload\new_file1.xlsx")
    print("Done")

