import streamlit as st
import pandas as pd
import openpyxl
import os
import zipfile
from CAL import perform
#from streamlit_extras.switch_page_button import switch_page

st.set_page_config(page_title="Revenue Forecasting", page_icon=":overview", layout="wide", initial_sidebar_state="collapsed")


def set_custom_styles():
    """
    Custom styles to hide Streamlit default elements and adjust margins.
    """

    custom_styles="""
    <style>
    MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility:hidden;}
    body {
        margin: 0;
        padding: 0;
        font-family: "Helvetica", sans-serif;
    }
    [data-testid="collapsedControl"] {
        display: none
    }     
    .main {
        # margin-left: -80px;
        padding: 20px;
        # margin-top:  -110px; 
        margin-left: -3rem;
        margin-top: -7rem;
        margin-right: -103rem;
    }
    .section-title {
        font-weight: bold;
        font-size: 24px;
    }
    .big-text {
        font-size: 20px;
    }

    .small-text {
        font-size: 14px;
    }

    </style>
    """
    st.markdown(custom_styles, unsafe_allow_html=True)
def custom_top_bar(selected_page=None):
    """
    Custom HTML for a fixed top bar.
    """
    selected_page = selected_page or "Report"
    # current_file = __file__.split("/")[-1]
    current_file = os.path.basename(__file__)
    custom_top_bar = f"""
    <style>
        #top-bar {{
            padding: 10px;
            # border-bottom: 1px solid #555;
            border-bottom: 1px solid #ccc;
            # color: black;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        #top-bar h3 {{
            padding: 10px;
            font-weight: bold;
            font-size: 24px;
        }}
        #nav-links {{
            display: flex;
            align-items: center;
        }}
        #nav-links a {{            
            text-decoration: none;
            margin-right: 20px;
            font-size: 16px;
            # font-family: "Source Sans Pro", sans-serif;
            # font-weight: 400;
            # transition: color 0.3s ease-in-out;
        }}
        #nav-links a:hover {{
            # color: #00bcd4;
            color: rgb(255, 75, 75);
        }}
    </style>
    <div id="top-bar">
        <h3 style="font-weight: bold; color: grey;">Revenue Forecasting</h3>
        <div id="nav-links">
            <a style="color: {'red' if selected_page == 'Home' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'Home' else 'none'}" href="/Home" target="_self">Home</a>
            <a style="color: {'red' if selected_page == 'Daily_Overview' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'Daily_Overview' else 'none'}" href="/Daily_Overview" target="_self">Daily Overview</a>
            <a style="color: {'red' if selected_page == 'Revenue_Analysis' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'Revenue_Analysis' else 'none'}" href="/Revenue_Analysis" target="_self">Revenue Analysis</a>
            <a style="color: {'red' if selected_page == 'Report' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'Report' else 'none'}" href="/Report" target="_self">Report</a>
            <a style="color: {'red' if selected_page == 'Upload' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'Upload' else 'none'}" href="/Upload" target="_self">Manage Collections</a>
            <a style="color: {'red' if selected_page == 'market' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'market' else 'none'}" href="/market" target="_self">market</a>
            <a style="color: {'red' if selected_page == 'Prediction' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'Prediction' else 'none'}" href="/Prediction" target="_self">Prediction</a>
            <a style="color: {'red' if selected_page == 'Trend' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'Trend' else 'none'}" href="/Trend" target="_self">trend</a>
            </div>
    </div>
    """
    st.markdown(custom_top_bar, unsafe_allow_html=True)

# custom_top_bar()
set_custom_styles()

url_path = st.experimental_get_query_params().get("pages", [""])[0]
url_to_page = {
    "/Home": "Home",
    "/Daily_Overview": "Daily_Overview",
    "/Revenue_Analysis": "Revenue_Analysis",
    "/Report": "Report",
   # "/Prediction": "Prediction",
    "/upload": "Upload",
    "/market": "market",
    "/Prediction": "Prediction",
    "/Trend": "Trend",
    #"/Yearly": "Yearly"
}
selected_page = url_to_page.get(url_path)
custom_top_bar(selected_page)
# -----------------------------------------------
UPLOAD_FOLDER = os.path.join(os.getcwd(), "Upload")
# st.set_page_config(page_title="Revenue Forecasting", page_icon=":barchart:", layout="wide")

# Create the upload folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# Add custom CSS styles
st.markdown(
    """
    <style>
    body {
        font-family: Arial, sans-serif;
        background-color: #f4f4f4;
        margin: 0;
        padding: 0;
    }
    .container {
        max-width: 800px;
        margin: 0 auto;
        padding: 20px;
        background-color: #ffffff;
        border-radius: 5px;
        box-shadow: 0 2px 5px rgba(0, 0, 0, 0.1);
    }
    .header {
        text-align: center;
        font-size: 24px;
        margin-bottom: 20px;
    }
    .input-section {
        margin-bottom: 20px;
    }
    .button-section {
        text-align: center;
    }
    .download-section {
        margin-top: 20px;
        text-align: center;
    }
    </style>
    """,
    unsafe_allow_html=True,
)
st.subheader('Forecast Report')

# User input section
st.markdown('<div class="input-section">', unsafe_allow_html=True)
col1, col2 = st.columns(2)
with col1:
    f1 = st.file_uploader("Upload File 1 (f1.xlsx)", type=["xlsx"])
with col2:
    f2 = st.file_uploader("Upload File 2 (f2.xlsx)", type=["xlsx"])
group_confirm = st.text_input("Group Confirm Number")
st.markdown('</div>', unsafe_allow_html=True)

# Process Data button
st.markdown('<div class="button-section">', unsafe_allow_html=True)
if st.button("Process Data"):
    if f1 and f2 and group_confirm:
        # Save uploaded files to the upload folder
        with open(os.path.join(UPLOAD_FOLDER, "f1.xlsx"), "wb") as f1_file:
            f1_file.write(f1.read())
        with open(os.path.join(UPLOAD_FOLDER, "f2.xlsx"), "wb") as f2_file:
            f2_file.write(f2.read())

        # # Load Excel files and sheets
        excel_file1 = openpyxl.load_workbook(os.path.join(UPLOAD_FOLDER, "f1.xlsx"))
        excel_sheet1 = excel_file1["Day on Day FC"]
        excel_sheet3 = excel_file1["Revenue Summary"]
        excel_sheet4 = excel_file1["Segment_Summary"]

        excel_file2 = openpyxl.load_workbook(os.path.join(UPLOAD_FOLDER, "f2.xlsx"))
        excel_sheet2 = excel_file2['History and Forecast Report']

        # Code for manually adding group confirm number
        excel_sheet2.cell(row=5, column=11).value = group_confirm
        perform(excel_file1,excel_sheet1, excel_sheet2,excel_sheet3,excel_sheet4)
        excel_file1.save(os.path.join(UPLOAD_FOLDER, "Final_Report.xlsx"))
        # Create a ZIP archive
        with zipfile.ZipFile(os.path.join(UPLOAD_FOLDER, "output.zip"), 'w') as zip_file:
            # Add each file to the archive
            for file_path in [os.path.join(UPLOAD_FOLDER, "f1.xlsx"), os.path.join(UPLOAD_FOLDER, "Final_Report.xlsx")]:
                zip_file.write(file_path)

        # Download the ZIP archive
        with open(os.path.join(UPLOAD_FOLDER, "output.zip"), "rb") as zip_file:
            zip_contents = zip_file.read()
        st.markdown("### Download Results")
        st.download_button("Download Output ZIP", data=zip_contents, file_name="output.zip")
st.markdown('</div>', unsafe_allow_html=True)

# Close the container div
st.markdown('</div>', unsafe_allow_html=True)

# Display instructions
# st.markdown('<div class="container">', unsafe_allow_html=True)
# st.write("Upload your files and enter the Group Confirm Number in the sidebar.")
st.markdown('</div>', unsafe_allow_html=True)