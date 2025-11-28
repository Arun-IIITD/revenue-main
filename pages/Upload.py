import streamlit as st
import pandas as pd
from pymongo import MongoClient
from datetime import datetime
from openpyxl import load_workbook
import io
import pymongo
import os
st.set_page_config(page_title="Revenue Forecasting", page_icon=":overview", layout="wide", initial_sidebar_state="collapsed")
def set_custom_styles():
    """
    Custom styles to hide Streamlit default elements and adjust margins.
    """

    custom_styles="""
    <style>
    MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility:hidden;}
    body {
        margin: 0;
        padding: 0;
        font-family: "Helvetica", sans-serif;
    }
    [data-testid="collapsedControl"] {
        display: none
    }     
    .main {
        # margin-left: -80px;
        padding: 20px;
        # margin-top:  -110px; 
        margin-left: -3rem;
        margin-top: -7rem;
        margin-right: -103rem;
    }
    .section-title {
        font-weight: bold;
        font-size: 24px;
    }
    .big-text {
        font-size: 20px;
    }

    .small-text {
        font-size: 14px;
    }

    </style>
    """
    st.markdown(custom_styles, unsafe_allow_html=True)
def custom_top_bar(selected_page=None):
    """
    Custom HTML for a fixed top bar.
    """
    selected_page = selected_page or "Upload"
    # current_file = __file__.split("/")[-1]
    current_file = os.path.basename(__file__)
    custom_top_bar = f"""
    <style>
        #top-bar {{
            padding: 10px;
            # border-bottom: 1px solid #555;
            border-bottom: 1px solid #ccc;
            # color: black;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        #top-bar h3 {{
            padding: 10px;
            font-weight: bold;
            font-size: 24px;
        }}
        #nav-links {{
            display: flex;
            align-items: center;
        }}
        #nav-links a {{            
            text-decoration: none;
            margin-right: 20px;
            font-size: 16px;
            # font-family: "Source Sans Pro", sans-serif;
            # font-weight: 400;
            # transition: color 0.3s ease-in-out;
        }}
        #nav-links a:hover {{
            # color: #00bcd4;
            color: rgb(255, 75, 75);
        }}
    </style>
    <div id="top-bar">
        <h3 style="font-weight: bold; color: grey;">Revenue Forecasting</h3>
        <div id="nav-links">
            <a style="color: {'red' if selected_page == 'Home' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'Home' else 'none'}" href="/Home" target="_self">Home</a>
            <a style="color: {'red' if selected_page == 'Daily_Overview' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'Daily_Overview' else 'none'}" href="/Daily_Overview" target="_self">Daily Overview</a>
            <a style="color: {'red' if selected_page == 'Revenue_Analysis' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'Revenue_Analysis' else 'none'}" href="/Revenue_Analysis" target="_self">Revenue Analysis</a>
            <a style="color: {'red' if selected_page == 'Report' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'Report' else 'none'}" href="/Report" target="_self">Report</a>
            <a style="color: {'red' if selected_page == 'Prediction' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'Prediction' else 'none'}" href="/Prediction" target="_self">Prediction</a>
            <a style="color: {'red' if selected_page == 'Upload' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'Upload' else 'none'}" href="/Upload" target="_self">Manage Collections</a>
            <a style="color: {'red' if selected_page == 'market' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'market' else 'none'}" href="/market" target="_self">market</a>
            <a style="color: {'red' if selected_page == 'Trend' else '#333'}; border-bottom: {'2px solid red' if selected_page == 'Trend' else 'none'}" href="/trend" target="_self">Trend</a>
            </div>
    </div>
    """
    st.markdown(custom_top_bar, unsafe_allow_html=True)

# custom_top_bar()
set_custom_styles()

url_path = st.experimental_get_query_params().get("pages", [""])[0]
url_to_page = {
    "/Home": "Home",
    "/Daily_Overview": "Daily_Overview",
    "/Revenue_Analysis": "Revenue_Analysis",
    "/Report": "Report",
    "/Prediction": "Prediction",
    "/upload": "Upload",
    "/market": "market",
    "/Trend": "Trend",
        }
selected_page = url_to_page.get(url_path)
custom_top_bar(selected_page)
#-------------------------------------------------------------------


st.markdown("\n")
st.markdown("\n")
connection_uri = "mongodb+srv://annu21312:6dPsrXPfhm19YxXl@hello.hes3iy5.mongodb.net/"
client = MongoClient(connection_uri, serverSelectionTimeoutMS=30000)
database_name = "Revenue_Forecasting"

# Define the db variable
db = client[database_name]
# Check password for specific collections
protected_collections = ["History_Fore", "Forecastin", "Covid", "History", "Prophet", "new"]


# Sidebar for configuration settings
st.subheader("Manage Forecasting Database")
col1, col2 = st.columns(2)
with col1:
    collection_name = st.text_input("Enter new/existing Collection Name:")

with col2:
    password = st.text_input("Enter Password(*For Admin Access):", type="password")
    if password == "AdityaRay117" :
        st.success("Password is correct. You have access to protected collections.")
uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx"])

# Create an empty DataFrame with specified columns
columns = ["Actual Date", "Business Date", "Rooms Sold", "Rooms for Sale", "Arrival Rooms", "Compliment Rooms",
           "House Use", "Hold", "Individual Confirm", "Individual Tentative", "Group Confirm", "Group Tentative",
           "Occupancy %", "Room Revenue", "ARR", "Inclusion Revenue", "Departure Rooms", "OOO Rooms", "Pax",
           "Individual Revenue", "Individual ARR", "Confirmed Group Revenue", "Confirmed Group ARR",
           "Tentative Group Revenue", "Tentative Group ARR", "Total Room Inventory"]
df = pd.DataFrame(columns=columns)
# st.title("Manage Collection")
with st.expander("**Existing Collections**", expanded=True):
    collections_in_db = client[database_name].list_collection_names()

    # Check if an Excel file is uploaded and a collection name is provided
    if uploaded_file and collection_name:
        try:
            # Display file details
            st.subheader("File Details:")
            st.info(f"File Name: {uploaded_file.name}")
            st.info(f"File Size: {uploaded_file.size} bytes")

            # Read the Excel file into a DataFrame
            wb = load_workbook(filename=io.BytesIO(uploaded_file.read()), read_only=True)
            ws = wb.active

            actual_date = datetime.now().strftime("%Y-%m-%d")

            # Iterate through all rows starting from the 3rd row
            for row in ws.iter_rows(min_row=5, values_only=True):
                if isinstance(row[0], (datetime, pd.Timestamp)):
                    business_date = row[0].strftime("%Y-%m-%d")

                    # Check if the business_date already exists in the collection
                    if db[collection_name].find_one({"Business Date": business_date}):
                        continue

                    row_dict = {"Actual Date": actual_date, "Business Date": business_date}

                    # Add the data from columns C to Z to the dictionary
                    for i in range(2, 26):
                        row_dict[columns[i]] = row[i]

                    # Append the row data to the dataframe
                    df = pd.concat([df, pd.DataFrame([row_dict])], ignore_index=True)
            st.warning(f"Duplicate rows discarded")

            # Drop duplicate rows based on the specified columns
            df.drop_duplicates(inplace=True)

            if df.empty:
                st.warning("No data to display.")
            else:
                st.subheader("Preview of the Uploaded Data:")
                st.dataframe(df)

            # Check if columns match the specified ones
            if set(df.columns) != set(columns):
                st.error("Column names do not match the specified ones.")
            else:
                # Check if the collection name is not empty
                if collection_name.strip() == "":
                    st.error("Collection name cannot be empty.")
                elif collection_name in protected_collections and password != "AdityaRay117":
                    st.error("Access denied. Upload button disabled for protected collections.")
                else:
                    # Button to upload data to the specified collection
                    if st.button("Upload", key="upload_button"):
                        with st.spinner("Uploading data..."):
                            # Add a timestamp to the data
                            df["Upload Timestamp"] = datetime.utcnow()

                            # Convert DataFrame to a list of dictionaries (JSON-like)
                            data_to_insert = df.to_dict(orient="records")

                            # Insert data into the specified collection
                            db[collection_name].insert_many(data_to_insert)

                        st.success("Data successfully uploaded!")

        except pymongo.errors.PyMongoError as mongo_error:
            st.error(f"Error with MongoDB: {mongo_error}")
        except Exception as e:
            st.error(f"Error processing data: {e}")

    # Display all collections in the database
    collections_in_db = client
    collections_in_db = client[database_name].list_collection_names()
    # st.subheader("All Collections in Database")

    if collection_name and (collection_name == "" or collection_name not in collections_in_db):
        st.info("No collection selected or the collection does not exist.")
    else:
        if collection_name and collection_name in collections_in_db:
            all_data = list(db[collection_name].find())
            # st.write(all_data)

            # Collection management section
            if all_data:
                all_df = pd.DataFrame(all_data)
                st.write(all_df)
                collections_in_db = client[database_name].list_collection_names()
                if collection_name in protected_collections and password != "AdityaRay117":
                    st.error("Only admins can make changes in this collection.")
                else:
                    st.markdown("**Delete Rows**")
                    # delete_rows_dates = st.multiselect("Choose Business Dates to Delete Rows:", all_df["Business Date"].unique())
                    multiselect_options = all_df.iloc[:, 2].unique()
                    delete_rows_dates = st.multiselect("Choose Business Dates to Delete Rows:", multiselect_options)

                    if st.button("Delete Rows"):
                        try:
                            # Delete rows based on the chosen Business Dates
                            db[collection_name].delete_many({"Business Date": {"$in": delete_rows_dates}})
                            st.success(f"Rows for selected Business Dates successfully deleted!")
                        except Exception as e:
                            st.error(f"Error deleting rows from database: {e}")

                    if collection_name not in protected_collections or password == "AdityaRay117":
                        st.markdown("**Drop Collection**")
                        st.markdown("This action will permanently delete this collection.")
                        if st.button("Drop Collection"):
                            try:
                                # Drop the entire collection
                                db[collection_name].drop()
                                st.success(f"Collection {collection_name} successfully dropped!")
                            except Exception as e:
                                st.error(f"Error dropping collection: {e}")
                    # else:
                    #     st.warning("This action will permanently delete the entire collection. Proceed with caution.")
            else:
                st.info("No data available in the collection.")
        else:
            # st.info("No collection selected or the collection does not exist.")

            collections_in_db = client[database_name].list_collection_names()
            if not collections_in_db:
                st.info("No collections found.")
            else:
                # st.write("**Collections:**")
                # st.markdown("<div style='font-size: 12px;'><strong>Collections:</strong></div>", unsafe_allow_html=True)

                for collection_name in collections_in_db:
                    # st.write(f"- {collection_name}")
                    st.markdown(f"<div style='font-size: 14px;'>- {collection_name}</div>", unsafe_allow_html=True)

